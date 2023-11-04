import cv2
import face_recognition
import os
import numpy as np
from datetime import datetime
import pickle
import openpyxl
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import automail

# Path to the directory containing student images
path = r'student_images'
attendee = []  # List of students present during attendance
images = []
classNames = []
mylist = os.listdir(path)

# Number of students (get from the number of rows in the student list file)
num_of_student = 0
wb = openpyxl.load_workbook("attendance/DiemDanh.xlsx")
sheetname = wb.sheetnames
sheet = wb[sheetname[0]]
num_of_student = sheet.max_row

# Get the current date
now = datetime.now()
date_str = now.strftime('%d-%B-%Y')

# File to store the number of attendance sessions
sobuoi_filename = 'attendance/sobuoi.txt'

# Read the current session number from the file or set it to 1
if not os.path.exists(sobuoi_filename):
    with open(sobuoi_filename, 'w') as f:
        f.write('1')

with open(sobuoi_filename, 'r') as f:
    sobuoi_str = f.read()
    if sobuoi_str.strip():
        sobuoi = int(sobuoi_str)
    else:
        sobuoi = 1

col_number = sobuoi

# Increment the session number and write it back to the file
with open(sobuoi_filename, 'w') as f:
    f.write(f'{sobuoi+1}')

# Clear the attendance and absent lists
with open('attendance/Attendance.csv', 'w') as f:
    f.write('')
with open('attendance/absent_list.txt', 'w') as f:
    f.write('')

# Load student images and their names
for cl in mylist:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])

# Function to encode faces in the images
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encoded_face = face_recognition.face_encodings(img)[0]
        encodeList.append(encoded_face)
    return encodeList

# Load the trained data if it exists, or train and save it
if os.path.exists('application_files\encoded_faces.pkl'):
    with open('application_files\encoded_faces.pkl', 'rb') as f:
        encoded_face_train = pickle.load(f)
else:
    encoded_face_train = findEncodings(images)
    with open('application_files\encoded_faces.pkl', 'wb') as f:
        pickle.dump(encoded_face_train, f)

# Function to mark attendance in the Excel sheet
def markAttendance(name):
    with open('attendance/Attendance.csv', 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if name not in nameList:
            now = datetime.now()
            time = now.strftime('%I:%M:%S:%p')
            date = now.strftime('%d-%B-%Y')
            f.writelines(f'{name}, {time}, {date}\n')
            attendee.append(name.upper())

# Function to write attendance to the Excel sheet
def write_to_excel():
    wb = openpyxl.load_workbook("attendance\DiemDanh.xlsx")
    sheetname = wb.sheetnames
    ws = wb[sheetname[0]]
    col_name = chr(ord('E') + col_number)
    cell_name = f"{col_name}1"
    ws[cell_name] = f"Attendance Session {col_number} - {date_str}"

    for i in range(2, num_of_student):
        if ws.cell(row=i, column=2).value in attendee:
            ws.cell(row=i, column=col_number + 5).value = True
        else:
            ws.cell(row=i, column=col_number + 5).value = False

    wb.save("attendance\DiemDanh.xlsx")

# Function to count the number of absences for each student
def count_absent():
    wb = openpyxl.load_workbook("attendance\DiemDanh.xlsx")
    sheetname = wb.sheetnames
    ws = wb[sheetname[0]]
    ws['Y1'] = "Number of Absences"
    for row in range(2, num_of_student):
        false_count = 0
        for col in range(6, 24):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == False:
                false_count += 1 
        ws.cell(row=row, column=25, value=false_count)
    wb.save("attendance\DiemDanh.xlsx")

# Function to create a file listing absent students
def create_absent_file():
    wb = openpyxl.load_workbook("attendance\DiemDanh.xlsx")
    sheetname = wb.sheetnames[0]
    ws = wb[sheetname]

    dsvang = []
    for row in range(2, num_of_student):
        vang = ws.cell(row=row, column=col_number+5).value
        if vang == False:
            msv = ws.cell(row=row, column=2).value
            ho = ws.cell(row=row, column=3).value 
            ten = ws.cell(row=row, column=4).value
            dsvang.append((msv, ho, ten)) 
    
    with open("attendance/absent_list.txt", "w", encoding="utf-8") as f:
        for msv, ho, ten in dsvang:
            f.write(f"{msv}, {ho} {ten}\n") 
    wb.close()    

# Initialize the webcam

def main():
    cap = cv2.VideoCapture(0)
    try:
        while True:
            success, img = cap.read()
            imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
            faces_in_frame = face_recognition.face_locations(imgS)
            encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
            for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                matches = face_recognition.compare_faces(
                    encoded_face_train, encode_face)
                faceDist = face_recognition.face_distance(
                    encoded_face_train, encode_face)
                matchIndex = np.argmin(faceDist)
                print(matchIndex)
                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    y1, x2, y2, x1 = faceloc
                    y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2-35), (x2, y2),
                                (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, name, (x1+6, y2-5),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    markAttendance(name.upper())
                else:
                    y1, x2, y2, x1 = faceloc
                    y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2-35), (x2, y2),
                                (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, 'UNKNOWN ', (x1+6, y2-5),
                                cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

                    
            cv2.imshow('Diem danh', img)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    except KeyboardInterrupt:
        pass
    write_to_excel()
    # Count absences, create absent file, and send emails
    count_absent()
    create_absent_file()
    time.sleep(1)
    print("Completed!")
    cap.release()
    cv2.destroyAllWindows()
if __name__ == '__main__':
    main()
