import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl

def send_emails():
        
    data_path = r'DiemDanh.xlsx'  
    index_path = r'index.html'   #email template , use $NAME to replace later
    sender_email = "duythong.ptit@gmail.com"  #Insert your email and password
    sender_password = "zjarmbbjvnxcybkd"    
    # Read email content from HTML file
    with open(index_path, "r", encoding="utf-8") as html_file:
        email_content = html_file.read()

    # Read data from Excel file
    wb = openpyxl.load_workbook(data_path)
    sheetname = wb.sheetnames
    ws = wb[sheetname[0]]
    ws.cell(row=1, column=24).value = "Da gui Email"
    recipients = []

    # Iterate through rows in the Excel sheet
    for row in range(2, ws.max_row + 1):
        false_count = 0

        # Check columns 6 to 24 for False values
        for col in range(6, 25):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == False:
                false_count += 1

        # If there are at least 2 False values and the "Check sent email" column is empty, add recipient
        if false_count >= 2 and ws.cell(row=row, column=24).value is None:
            ho = ws.cell(row=row, column=3).value
            ten = ws.cell(row=row, column=4).value
            full_name = f"{ho} {ten}"
            email = ws.cell(row=row, column=5).value
            if email:
                recipients.append((full_name, email))
                ws.cell(row=row, column=24).value = True

    # Save the workbook with the updates
    wb.save(data_path)

    # Write recipient names and emails to a text file
    with open("sent_emails.txt", "w+", encoding="utf-8") as f:
        for recipient in recipients:
            f.write(f"{recipient[0]}, {recipient[1]}\n")

    # Connect to the SMTP server of Gmail
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()

    # Log in to the email account
    server.login(sender_email, sender_password)

    # Send emails to each recipient
    subject = "[CẢNH BÁO NGHỈ QUÁ 2 BUỔI HỌC] "
    for full_name, receiver_email in recipients:
        # Create a MIMEMultipart object to compose the email
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject

        # Replace "$NAME" in the email content with the recipient's full name
        email_content_with_name = email_content.replace("$NAME", full_name)

        # Attach the HTML email content
        message.attach(MIMEText(email_content_with_name, "html"))

        # Send the email
        server.sendmail(sender_email, receiver_email, message.as_string())

    # Close the SMTP connection
    server.quit()
    print("Emails have been sent successfully!")

